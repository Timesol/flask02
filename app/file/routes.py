import os
from os import listdir
from os.path import isfile, join
from flask import send_file
import pandas as pd
from app.pandaex import  sendpandas
from app.file import bp
from flask_login import login_required
import requests
from requests import Request, Session
from flask import render_template, flash, redirect, url_for, request, g, current_app, json
from flask_login import login_user, logout_user, current_user, login_required
from flask_babel import _, get_locale
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl import load_workbook
from app.pandaex import  sendpandas
from app.file.forms import SearchForm

@bp.before_app_request
def before_request():
    if current_user.is_authenticated:
        
        g.search_form = SearchForm()
    g.locale = str(get_locale())


UPLOAD_FOLDER = os.environ.get('UPLOAD_FOLDER_ENV')
UPLOAD_FOLDER_CONFIGS = os.environ.get('UPLOAD_FOLDER_CONFIGS')
ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'xlsx','cfg'])
ALLOWED_EXTENSIONS_PANDAS = set(['xlsx'])
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_file_pandas(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_PANDAS
    return True  






@bp.route('/uploads', methods=['GET', 'POST'])
@login_required
def uploads():
    list=[]
    for x in os.listdir(UPLOAD_FOLDER):
        list.append(x)
    print(list)

        
    

    if request.method == 'POST':
        # check if the post request has the file part
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        # if user does not select file, browser also
        # submit a empty part without filename
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_FOLDER, filename))
            flash(_('File uploaded!'))
            #if allowed_file_pandas(file.filename):
                #sendpandas(filename)
                
          

            
        return redirect(url_for('file.uploads',
                            filename=filename, ))
       

    
    return render_template('uploads.html', title=_('Uploads'), list=list)


@bp.route('/return_files/<filename>')
def return_files(filename):
    try:
        return send_file(UPLOAD_FOLDER+filename , attachment_filename=filename)
    except Exception as e:
        return str(e)
    return render_template('uploads.html', title=_('Uploads'), list=list)






@bp.route('/expy',methods=['GET', 'POST'])
@login_required




def expy(cat,scat,hard,user,tech,cust,contr,time,date):
    
    folder=os.environ.get('EXCEL_FOLDER_ENV')+current_user.username+'/'+'statistic_'+date+"_"+current_user.username+'.xlsx'
    
    sheet=cat
    print(cat)
    srcfile = openpyxl.load_workbook(folder,read_only=False)
    sheetname = srcfile.get_sheet_by_name(sheet)
    save("A",tech, sheetname)
    save("B",cust, sheetname)
    save("C",contr, sheetname)
    save("D",hard, sheetname)
    save("E",time, sheetname)
    save("F",user, sheetname)
    save("G",scat, sheetname)


    srcfile.save(folder)


def expynew(date):
    
    folder=os.environ.get('EXCEL_FOLDER_ENV')+current_user.username+'/'+'statistic_'+date+"_"+current_user.username+'.xlsx'
    
   
    empty_row=0
    
    wb = openpyxl.Workbook()
    
    wb.save(folder)
    sheetname_original = wb.get_sheet_by_name('Sheet')

    sheetname_original.title='Review'
    review=wb.get_sheet_by_name('Review')
    
    wb.save(folder)
    wb.create_sheet('Installation')
    wb.create_sheet('Commissioning')
    wb.create_sheet('Dismantling')
    wb.create_sheet('CIA')
    wb.create_sheet('MPLS')
    wb.create_sheet('Disruption')
    wb.save(folder)
    
    data=[('Technology', 'Customer','Contract','Hardware','Time','Employee','Subcategory','Additional Info')]
    for isheet in wb.sheetnames:

        if isheet == 'Review':
            continue
        
        isheet = wb.get_sheet_by_name(isheet)
        for row in data:
           isheet.append(row)
        

    wb.save(folder)

    
    
    


    


    

    



def save(l, t,sheetname):
    empty_row=1
    
    for cell in sheetname[l]:
        
        if cell.value is None:
            break
        empty_row =(cell.row)+1
            

         
    if not t:
        print('Object is none')

        sheetname[l+"%d"  % empty_row]= str("/")
        return True
    sheetname[l+"%d"  % empty_row]= str(t)



def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print ('Error: Creating directory. ' +  directory)



@bp.route('/return_files_user/<filename>')
def return_files_user(filename):
    try:
        return send_file(UPLOAD_FOLDER+current_user.username+'/'+filename , attachment_filename=filename)
    except Exception as e:
        return str(e)
    return render_template('user.html', title=_('user'), filelist=filelist)


@bp.route('/return_files_router/<filename>')
def return_files_router(filename):
    try:
        return send_file(UPLOAD_FOLDER_CONFIGS+filename , attachment_filename=filename)
    except Exception as e:
        return str(e)

